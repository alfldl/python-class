# 파티 참가 대상  User(member.xlsx)들에게 자동으로 Mail 보내기
from email.message import EmailMessage
import smtplib
import openpyxl

def check_party_members():
    # Excel Open
    wb = openpyxl.load_workbook('C:\\PyCharmProject\\Sources\\Basic\\ch27\\member.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    party_members={}
    # 2 Row 부터 max_row 까지 진행
    for r in range(2, sheet.max_row + 1):
        payment = sheet.cell(row=r, column=sheet.max_column).value
        if payment == 'invite':
            name = sheet.cell(row=r, column=1).value
            email = sheet.cell(row=r, column=2).value
            party_members[name] = email
    return party_members

def main():
    #  파티 참가 명단 확보
    party_members = check_party_members()
    print('main  party members  {}...'.format(party_members))
    # Log In eMail account
    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    # smtp 정상여부 Hello Test
    smtp.ehlo()
    # smtp Security  적용
    smtp.starttls()
    # smtp Login
    smtp.login('ttaekwang3@gmail.com', 'ktg9489#@!')

    # 파티 참가 대상 회원에게 mail을 보냄
    for name , email in party_members.items():
        # Message 생성
        msg = EmailMessage()

        # 내부적으로 dic으로 선언
        msg['Subject'] = '회원 파티 초대장(Mail Test)'
        msg['From'] = 'ttaekwang3@gmail.com'
        msg['to'] = email
        msg.set_content('''안녕하세요 {}회원님 ,
                         2019년 09월 저녁파티에 참석요청합니다
                         빠른 시간내에 회신 부탁드려요
                          '''.format(name))
        # Mail 전송
        print('Sending mail to {}...'.format(email))
        sendmail_status = smtp.send_message(msg)

        if sendmail_status != {}:
            print('{}에게 메일 전송이 실패하였습니다 {}...'.format(email, sendmail_status))
        else:
            print('{}에게 메일 전송이 성공하였습니다 {}...'.format(email, sendmail_status))


    # 연결 종료
    smtp.quit()


if __name__ == '__main__':
    main()
